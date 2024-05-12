import streamlit as st
import pandas as pd
import os
import pyrebase
import time
import base64
import hashlib
from base64 import b64encode
import openpyxl

file_paths_with_titles = {
    "TIMBAO-BUKAL-BUKAL CIS.xlsx": "TIMBAO-BUKAL",
    "2 row.xlsx": "DILA",
    "Ma. Pelaez.xlsx": "Ma. Pelaez",
    "Puypuy.xlsx": "Puypuy",
    "Bangyas.xlsx": "Bangyas",
    "CANLUBANG CIS.xlsx": "CANLUBANG",
    "PEREZ.xlsx": "PEREZ",
    "LAMOT I.xlsx": "LAMOT I",
    "LAMOT II.xlsx": "LAMOT II",
    "MASIIT CIS.xlsx": "MASIIT CIS",
    "PRINZA.xlsx": "PRINZA",
    "ARJONA CIS.xlsx": "ARJONA CIS",
    "CALUMPANG CIS.xlsx": "CALUMPANG CIS",
    "TUY-BAANAN.xlsx": "TUY-BAANAN",
    "BUNGKOL CIS.xlsx": "BUNGKOL CIS",
    "MASIKAP CIS.xlsx": "MASIKAP CIS",
    "BANAGO CIS.xlsx": "BANAGO CIS",
    "BANILAD CIS.xlsx": "BANILAD CIS",
    "PALAYAN CIS.xlsx": "PALAYAN CIS",
    "TAYTAY CIS.xlsx": "TAYTAY CIS",
    "NAGCALBANG PCIS.xlsx": "NAGCALBANG PCIS",
    "LAGUAN CIS.xlsx": "LAGUAN CIS",
    "TALAGA CIS.xlsx": "TALAGA CIS",
    "TALA CIS.xlsx": "TALA CIS",
    "MAYTON CIS.xlsx": "MAYTON CIS",
    "BALANGA CIS.xlsx": "BALANGA CIS",
    "BANADERO CIS.xlsx": "BANADERO CIS",
    "STO. ANGEL CIS.xlsx": "STO. ANGEL CIS",
    "Sta. Veronica CIS.xlsx": "Sta. Veronica CIS",
    "Sta. Isabel CIS.xlsx": "Sta. Isabel CIS",
    "San Benito CIS.xlsx": "San Benito CIS",
    "San Roque PCIS.xlsx": "San Roque PCIS",
    "Cavinti CIS.xlsx": "Cavinti CIS",
    "Sumucab CIS.xlsx": "Sumucab CIS",
    "Lilian CIS.xlsx": "Lilian CIS",
    "Salang De Castro CIS.xlsx": "Salang De Castro CIS",
    "LONGOS CIS.xlsx": "LONGOS CIS",
    "San Antonio CIS.xlsx": "San Antonio CIS",
    "KALAYAAN CIS.xlsx": "KALAYAAN CIS",
    "ILOG-KAWAYAN CIS.xlsx": "ILOG-KAWAYAN CIS",
    "Binambang PCIS.xlsx": "Binambang PCIS",
    "Concepcion CIS.xlsx": "Concepcion CIS",
    "WAWA IBAYO CIS.xlsx": "WAWA IBAYO CIS",
    "PAAGAHAN CIS.xlsx": "PAAGAHAN CIS",
    "MATALA-TALA CIS.xlsx": "MATALA-TALA CIS",
    "MARAVILLA CIS.xlsx": "MARAVILLA CIS",
    "MAIMPEZ CIS.xlsx": "MAIMPEZ CIS",
    "BAYUCAIN CIS.xlsx": "BAYUCAIN CIS",
    "BUKAL CIS.xlsx": "BUKAL CIS",
    "GAGALOT CIS.xlsx": "GAGALOT CIS",
    "San Roque CIS.xlsx": "San Roque CIS",
    "Sta. Catalina CIS.xlsx": "Sta. Catalina CIS",
    "Amonoy CIS.xlsx": "Amonoy CIS",
    "MALINAO CIS.xlsx": "MALINAO CIS",
    "Panglan CIS.xlsx": "Panglan CIS",
    "PANGIL-CORALAO-TALORTOR CIS.xlsx": "PANGIL-CORALAO-TALORTOR CIS",
    "PAETA CIS.xlsx": "PAETA CIS",
    "TICDAO CIS.xlsx": "TICDAO CIS",
    "BANILAN CIS.xlsx": "BANILAN CIS",
    "MATIKIW-1 CIS.xlsx": "MATIKIW-1 CIS",
    "MATIKIW-2 CIS.xlsx": "MATIKIW-2 CIS",
    "CASA REAL CIS.xlsx": "CASA REAL CIS",
    "KABULUSAN CIS.xlsx": "KABULUSAN CIS",
    "BALIAN TAVERA CIS.xlsx": "BALIAN TAVERA CIS",
    "BOLLERO BALIAN CIS.xlsx": "BOLLERO BALIAN CIS",
    "SULIB CIS.xlsx": "SULIB CIS",
    "SAN JOSE CIS.xlsx": "SAN JOSE CIS",
    "CALANGAY CIS.xlsx": "CALANGAY CIS",
    "LSPU SIS.xlsx": "LSPU SIS",
    "ROMELO PINAIT CIS.xlsx": "ROMELO PINAIT CIS",
    "WAWA PCIS.xlsx": "WAWA PCIS",
    "MAKATAD PCIS.xlsx": "MAKATAD PCIS",
    "MAPAGONG PCIS.xlsx": "MAPAGONG PCIS",
    "LAMAO-LATI CIS.xlsx": "LAMAO-LATI CIS",
    "LAMBAC PCIS.xlsx": "LAMBAC PCIS"
}



# Dictionary mapping town names to file paths
town_file_paths = {
    "ARJONA COMMUNAL IRRIGATION SYSTEM": "new data/ARJONA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BALANGA COMMUNAL IRRIGATION SYSTEM": "new data/BALANGA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BALIAN-TAVERA COMMUNAL IRRIGATION SYSTEM": "new data/BALIAN-TAVERA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BAÑADERO COMMUNAL IRRIGATION SYSTEM": "new data/BAÑADERO COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BANAGO COMMUNAL IRRIGATION SYSTEM": "new data/BANAGO COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BANGYAS COMMUNAL IRRIGATION SYSTEM": "new data/BANGYAS COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BANILAD COMMUNAL IRRIGATION SYSTEM": "new data/BANILAD COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BANILAN COMMUNAL IRRIGATION SYSTEM": "new data/BANILAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "CALANGAY COMMUNAL IRRIGATION SYSTEM": "new data/CALANGAY COMMUNAL IRRIGATION SYSTEM.xlsx",
    "DILA COMMUNAL IRRIGATION SYSTEM": "new data/DILA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "ILOG KAWAYAN COMMUNAL IRRIGATION SYSTEM": "new data/ILOG KAWAYAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "LAGUAN COMMUNAL IRRIGATION SYSTEM": "new data/LAGUAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BAYUCAIN COMMUNAL IRRIGATION SYSTEM": "new data/BAYUCAIN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "BOLLERO-BALIANCOMMUNAL IRRIGATION SYSTEM": "new data/BOLLERO-BALIANCOMMUNAL IRRIGATION SYSTEM.xlsx",
    "BUKAL COMMUNAL IRRIGATION SYSTEM": "new data/BUKAL COMMUNAL IRRIGATION SYSTEM.xlsx",
    "CALUMPANG COMMUNAL IRRIGATION SYSTEM": "new data/CALUMPANG COMMUNAL IRRIGATION SYSTEM.xlsx",
    "CASA REAL COMMUNAL IRRIGATION SYSTEM": "new data/CASA REAL COMMUNAL IRRIGATION SYSTEM.xlsx",
    "CAVINTI COMMUNAL IRRIGATION SYSTEM": "new data/CAVINTI COMMUNAL IRRIGATION SYSTEM.xlsx",
    "DAYAP COMMUNAL IRRIGATION SYSTEM": "new data/DAYAP COMMUNAL IRRIGATION SYSTEM.xlsx",
    "LAMOT-I COMMUNAL IRRIGATION SYSTEM": "new data/LAMOT-I COMMUNAL IRRIGATION SYSTEM.xlsx",
    "LAMOT-2 COMMUNAL IRRIGATION SYSTEM": "new data/LAMOT-2 COMMUNAL IRRIGATION SYSTEM.xlsx",
    "LILIAN COMMUNAL IRRIGATION SYSTEM": "new data/LILIAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "LSPU SMALL IRRIGATION SYSTEM": "new data/LSPU SMALL IRRIGATION SYSTEM.xlsx",
    "MA. PELAEZ COMMUNAL IRRIGATION SYSTEM": "new data/MA. PELAEZ COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MAIMPEZ COMMUNAL IRRIGATION SYSTEM": "new data/MAIMPEZ COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MAKATAD COMMUNAL IRRIGATION SYSTEM": "new data/MAKATAD COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MAPAGONG COMMUNAL IRRIGATION SYSTEM": "new data/MAPAGONG COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MARAVILLA COMMUNAL IRRIGATION SYSTEM": "new data/MARAVILLA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MASIIT COMMUNAL IRRIGATION SYSTEM": "new data/MASIIT COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MATALATALA COMMUNAL IRRIGATION SYSTEM": "new data/MATALATALA COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MATIKIW-1 COMMUNAL IRRIGATION SYSTEM": "new data/MATIKIW-1 COMMUNAL IRRIGATION SYSTEM.xlsx",
    "MATIKIW-2 COMMUNAL IRRIGATION SYSTEM": "new data/MATIKIW-2 COMMUNAL IRRIGATION SYSTEM.xlsx",
    "PAAGAHAN COMMUNAL IRRIGATION SYSTEM": "new data/PAAGAHAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "PALAYAN COMMUNAL IRRIGATION SYSTEM": "new data/PALAYAN COMMUNAL IRRIGATION SYSTEM.xlsx",
    "PEREZ COMMUNAL IRRIGATION SYSTEM": "new data/PEREZ COMMUNAL IRRIGATION SYSTEM.xlsx",
    "SAN ANTONIO COMMUNAL IRRIGATION SYSTEM": "new data/SAN ANTONIO COMMUNAL IRRIGATION SYSTEM.xlsx",
    "SAN BENITO COMMUNAL IRRIGATION SYSTEM": "new data/SAN BENITO COMMUNAL IRRIGATION SYSTEM.xlsx",
    "SAN JOSE COMMUNAL IRRIGATION SYSTEM": "new data/SAN JOSE COMMUNAL IRRIGATION SYSTEM.xlsx",
    "SAN ROQUE COMMUNAL IRRIGATION SYSTEM": "new data/SAN ROQUE COMMUNAL IRRIGATION SYSTEM.xlsx",
    "STA. CATALINA COMMUNAL IRRIGATION SYSTEM": "new data/STA. CATALINA COMMUNAL IRRIGATION SYSTEM.xlsx"
    
}

# Define a list of file paths
file_paths = [
    "TIMBAO-BUKAL-BUKAL CIS.xlsx",
    "2 row.xlsx",
    "Ma. Pelaez.xlsx",
    "Puypuy.xlsx",
    "Bangyas.xlsx",
    "PEREZ.xlsx",
    "LAMOT I.xlsx",
    "LAMOT II.xlsx",
    "CANLUBANG CIS.xlsx",
    "MASIIT CIS.xlsx",
    "PRINZA.xlsx",
    "ARJONA CIS.xlsx",
    "CALUMPANG CIS.xlsx",
    "TUY-BAANAN.xlsx",
    "BUNGKOL CIS.xlsx",
    "MASIKAP CIS.xlsx",
    "BANAGO CIS.xlsx",
    "BANILAD CIS.xlsx",
    "PALAYAN CIS.xlsx",
    "TAYTAY CIS.xlsx",
    "NAGCALBANG PCIS.xlsx",
    "LAGUAN CIS.xlsx",
    "TALAGA CIS.xlsx",
    "TALA CIS.xlsx",
    "MAYTON CIS.xlsx",
    "BALANGA CIS.xlsx",
    "BANADERO CIS.xlsx",
    "STO. ANGEL CIS.xlsx",
    "Sta. Veronica CIS.xlsx",
    "Sta. Isabel CIS.xlsx",
    "San Benito CIS.xlsx",
    "San Roque PCIS.xlsx",
    "Cavinti CIS.xlsx",
    "Sumucab CIS.xlsx",
    "Lilian CIS.xlsx",
    "Salang De Castro CIS.xlsx",
    "LONGOS CIS.xlsx",
    "San Antonio CIS.xlsx",
    "KALAYAAN CIS.xlsx",
    "ILOG-KAWAYAN CIS.xlsx",
    "Binambang PCIS.xlsx",
    "Concepcion CIS.xlsx",
    "WAWA IBAYO CIS.xlsx",
    "PAAGAHAN CIS.xlsx",
    "MATALA-TALA CIS.xlsx",
    "MARAVILLA CIS.xlsx",
    "MAIMPEZ CIS.xlsx",
    "BAYUCAIN CIS.xlsx",
    "BUKAL CIS.xlsx",
    "GAGALOT CIS.xlsx",
    "San Roque CIS.xlsx",
    "Sta. Catalina CIS.xlsx",
    "Amonoy CIS.xlsx",
    "MALINAO CIS.xlsx",
    "Panglan CIS.xlsx",
    "PANGIL-CORALAO-TALORTOR CIS.xlsx",
    "PAETA CIS.xlsx",
    "TICDAO CIS.xlsx",
    "BANILAN CIS.xlsx",
    "MATIKIW-1 CIS.xlsx",
    "MATIKIW-2 CIS.xlsx",
    "CASA REAL CIS.xlsx",
    "KABULUSAN CIS.xlsx",
    "BALIAN TAVERA CIS.xlsx",
    "BOLLERO BALIAN CIS.xlsx",
    "SULIB CIS.xlsx",
    "SAN JOSE CIS.xlsx",
    "CALANGAY CIS.xlsx",
    "LSPU SIS.xlsx",
    "ROMELO PINAIT CIS.xlsx",
    "WAWA PCIS.xlsx",
    "MAKATAD PCIS.xlsx",
    "MAPAGONG PCIS.xlsx",
    "LAMAO-LATI CIS.xlsx",
    "LAMBAC PCIS.xlsx"
]

num_files = len(file_paths)

# Function to read and display the Excel file based on the current index
def display_inventory(file_index):
    # Read the Excel file
    df_inventory = pd.read_excel(file_paths[file_index])
    title = file_paths_with_titles.get(file_paths[file_index], "Unknown Title")
    st.markdown(f"## {title}")
    # Display the contents of the Excel file
    st.write(df_inventory)

def get_img_as_base64(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode('utf-8')    

def set_background(background_image_path):
    background_image_ext = 'png'  # Modify the extension if needed
    encoded_image = base64.b64encode(open(background_image_path, "rb").read()).decode()
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url('data:image/{background_image_ext};base64,{encoded_image}');
            background-size: cover;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

def search_inventory(search_query):
    # Clear the previous output
    st.session_state.search_query = search_query
    st.session_state.search_result_found = False
    st.session_state.search_result_container = None

    # Clear the sidebar background image
    st.sidebar.markdown("")
    
    # Clear the previous inventory display
    st.markdown("")

    # Initialize a variable to track if any search result is found
    search_result_found = False

    # Check if the search query is empty
    if not search_query:
        st.warning("Please enter a search query.")
        return

    # Search for the title in the file_paths_with_titles dictionary
    for file_path, title in file_paths_with_titles.items():
        if search_query.lower() in title.lower():
            # If the title matches the search query, display the data of the corresponding file
            df_inventory = pd.read_excel(file_path)
            result_container = st.container()
            st.session_state.search_result_container = result_container
            with result_container:
                st.markdown(f"## {title}")
                st.write(df_inventory)
            search_result_found = True

    # Update the session state with the search result status
    st.session_state.search_result_found = search_result_found

    # If no matching title is found, display a message
    if not search_result_found:
        st.warning("No matching title found.")

def sidebar_bg(side_bg):
    try:
        # Get the file extension
        _, side_bg_ext = os.path.splitext(side_bg)
        # Read the image file and encode it as base64
        with open(side_bg, "rb") as img_file:
            img_data = img_file.read()
            encoded_img = base64.b64encode(img_data).decode()
        # Set the background style for the sidebar
        st.markdown(
            f"""
            <style>
            [data-testid="stSidebar"] > div:first-child {{
                background: url(data:image/{side_bg_ext};base64,{encoded_img}) !important;
            }}
            </style>
            """,
            unsafe_allow_html=True,
        )
    except FileNotFoundError:
        st.error(f"Image file '{side_bg}' not found.")
    except Exception as e:
        st.error(f"Error: {e}")

side_bg = "static/image1.jpg"
sidebar_bg(side_bg)


# Firebase initialization
firebaseConfig = {
    'apiKey': "AIzaSyAF25avieiXm9XIZezYaL1JCEAux0-Gl1w",
    'authDomain': "nia-mis.firebaseapp.com",
    'databaseURL': 'https://nia-mis-default-rtdb.firebaseio.com/',
    'projectId': "nia-mis",
    'storageBucket': "nia-mis.appspot.com",
    'messagingSenderId': "333239771011",
    'appId': "1:333239771011:web:51f1310ff4894b6cc0e0b6",
    'measurementId': "G-SND6D820B7"
}


firebase = pyrebase.initialize_app(firebaseConfig)
auth = firebase.auth()
db = firebase.database()

def authenticate(email, password):
    try:
        # Check if the email is in declined accounts
        declined_accounts = db.child("declined_accounts").get().val()
        if declined_accounts and email in declined_accounts:
            st.error("Your account request has been declined.")
            return None

        # Check if the email/password match any user in approved_users
        approved_users = db.child("approved_users").get().val()
        if approved_users:
            for user_email, user_data in approved_users.items():
                if email == user_data.get("email") and password == user_data.get("password"):
                    # Return the authenticated user data
                    return user_data

        # If not found in approved_users or declined_accounts, attempt authentication using Firebase Authentication
        login = auth.sign_in_with_email_and_password(email, password)
        return login
    except Exception as e:
        error_message = str(e)
        if "INVALID_PASSWORD" in error_message or "INVALID_EMAIL" in error_message:
            st.error("Invalid email or password.")
        else:
            st.error("Wrong email or password.")
        return None
    
# Function to send a password reset email
def send_password_reset_email(email):
    if not email:
        st.toast("Please, input email.")
        time.sleep(2)
        return

    try:
        auth.send_password_reset_email(email)
        st.toast("Password reset email sent.")
        time.sleep(2)
    except Exception as e:
        error_message = str(e)
        if "INVALID_EMAIL" in error_message:
            st.toast("Invalid email address. Please check your email.")
            time.sleep(2)
        elif "MISSING_EMAIL" in error_message:
            st.toast("Missing email address. Please enter your email.")
            time.sleep(2)
        else:
            st.toast(f"Error sending password reset email: {e}")
            time.sleep(2)


def signup(email, password):
    try:
        # Store user details in pending sign-up requests
        db.child("pending_signups").child(email.replace(".", ",")).set({"email": email, "password": password})
        st.success("Sign up successful. Please wait for admin approval.")
    except Exception as e:
        st.error(f"Error signing up: {e}")

def validate_email(email):
    if "@" not in email or "." not in email:
        return False
    return True

def validate_password(password):
    if len(password) < 6:
        return False
    return True

def approve_signup(email):
    try:
        # Get the sign-up request data
        signup_data = db.child("pending_signups").child(email.replace(".", ",")).get().val()
        if signup_data:
            # Move user to approved users list
            db.child("approved_users").child(email.replace(".", ",")).set(signup_data)
            # Delete user from pending sign-ups
            db.child("pending_signups").child(email.replace(".", ",")).remove()
            st.success(f"Account for {email} approved.")
        else:
            st.error("No sign-up request found for the given email.")
    except Exception as e:
        st.error(f"Error approving sign-up request: {e}")

# Function to get pending sign-up requests
def get_pending_signups():
    try:
        pending_signups = db.child("pending_signups").get().val()
        return pending_signups
    except Exception as e:
        st.error(f"Error retrieving pending sign-up requests: {e}")
        return None
    
def get_firebase_db_users():
    try:
        # Retrieve list of users from Firebase Realtime Database
        user_data = db.child("approved_users").get().val()
        if user_data:
            user_emails = [data.get("email") for data in user_data.values()]  # Accessing email using get method
            return user_emails
        else:
            return []
    except Exception as e:
        st.error(f"Error retrieving Firebase Realtime Database users: {e}")
        return []
    
def change_password(email, old_password, new_password):
    try:
        # Authenticate the user using the email and old password from the Realtime Database
        user_data = db.child("approved_users").order_by_child("email").equal_to(email).get()
        if user_data.each():
            for user in user_data.each():
                user_email = user.val().get("email")
                user_password = user.val().get("password")
                if user_email == email and user_password == old_password:
                    # Change password
                    db.child("approved_users").child(user.key()).update({"password": new_password})
                    st.success("Password changed successfully.")
                    return
            # If the email or old password doesn't match
            st.error("Invalid email or old password.")
        else:
            # If the email is not found
            st.error("User not found.")
    except Exception as e:
        st.error(f"Error changing password: {e}")
    
def manage_accounts():
    try:
        # Dropdown menu for account actions
        action = st.selectbox("Select Action:", ["Approval", "Account Deletion", "Change Password"])

        if action == "Approval":
            st.subheader("Approval of Account")
            pending_signups = get_pending_signups()

            if pending_signups:
                st.write("Pending Sign-up Requests:")
                for email, signup_data in pending_signups.items():
                    st.write(f"Email: {email}")
                    approve_button = st.button(f"Approve {email}")
                    decline_button = st.button(f"Decline {email}")
                    if approve_button:
                        approve_signup(email)
                    elif decline_button:
                        decline_signup(email)
            else:
                st.write("No pending sign-up requests.")

        elif action == "Account Deletion":
            st.subheader("Delete Account")
            firebase_db_users = get_firebase_db_users()
            
            email_to_delete = st.selectbox("Select Email Address to Delete:", [""] + firebase_db_users)
            
            if st.button("Delete Account") and email_to_delete:
                delete_account(email_to_delete)

        elif action == "Change Password":
            st.subheader("Change Password")
            email = st.text_input("Email Address")
            old_password = st.text_input("Old Password", type="password")
            new_password = st.text_input("New Password", type="password")
            confirm_password = st.text_input("Confirm New Password", type="password")
            
            if st.button("Change Password"):
                if email and old_password and new_password and confirm_password:
                    if new_password == confirm_password:
                        change_password(email, old_password, new_password)
                    else:
                        st.error("New passwords do not match.")
                else:
                    st.error("Please fill in all fields.")

    except Exception as e:
        st.error(f"Error managing accounts: {e}")

def decline_signup(email):
    try:
        # Remove user details from pending sign-up requests
        db.child("pending_signups").child(email.replace(".", ",")).remove()
        st.success(f"Request from {email} has been declined.")
    except Exception as e:
        st.error(f"Error declining sign-up request: {e}")

def delete_account(email):
    try:
        # Check if the account exists in approved_users
        user_ref = db.child("approved_users").order_by_child("email").equal_to(email).get()
        if user_ref.each():
            # Remove the user from approved_users
            for user in user_ref.each():
                db.child("approved_users").child(user.key()).remove()
            st.success(f"Account for {email} deleted successfully.")
        else:
            st.error("No account found for the given email.")
    except Exception as e:
        st.error(f"Error deleting account: {e}")


# Function to logout user
def logout():
    # Clear user info from session state
    st.session_state.user = None


def save_to_csv(new_entry, csv_file_path):
    try:
        # Check if the CSV file already exists
        if os.path.exists(csv_file_path):
            # Read the existing CSV file
            df = pd.read_csv(csv_file_path, encoding='latin1')
            
            # Check if the new entry already exists in the DataFrame
            if df.equals(pd.DataFrame([new_entry])):
                st.warning("Entry already exists.")
                return
            
            # Append the new entry to the DataFrame
            df = df.append(new_entry, ignore_index=True)
        else:
            # If the file doesn't exist, create a new DataFrame with the new entry
            df = pd.DataFrame([new_entry])

        # Save the DataFrame to CSV
        df.to_csv(csv_file_path, index=False, encoding='latin1')
        st.success("Entry saved successfully!")
    except Exception as e:
        st.error(f"Error saving entry: {e}")


def edit_entry(index, edited_entry, csv_file_path):
    try:
        # Read the CSV file
        df = pd.read_csv(csv_file_path, encoding='latin1')

        # Update the entry at the specified index
        df.loc[index] = edited_entry

        # Save the DataFrame back to CSV
        df.to_csv(csv_file_path, index=False)
        st.success("Entry edited successfully!")
    except Exception as e:
        st.error(f"Error editing entry: {e}")

# Streamlit app content
def main():
    background_image_path = "static/2nd.png"  # Adjust the path accordingly
    global csv_file_path  # Declare csv_file_path as a global variable
    csv_file_path = r'sample.csv'

    if 'file_index' not in st.session_state:
        st.session_state.file_index = 0

    # Create a session state object
    if 'user' not in st.session_state:
        st.session_state.user = None  # Initialize st.session_state.user to None


    if st.session_state.user is None:
        st.markdown("<h1 style=' color: #545454;'>LOGIN</h1>", unsafe_allow_html=True)

        st.markdown(
        """
        <style>
        .st-emotion-cache-q8sbsg p {
            color: black;
        }
        .st-emotion-cache-16idsys p{
            color: #545454;
        }
        button.st-emotion-cache-hc3laj.ef3psqc12 {
            background-color: #2ECC71;
            position: relative;
            border: 1px solid black;
            margin: 0;
            color: #fff;
            display: inline-block;
            text-decoration: none;
            text-align: center;
        }
        .st-b2 {
        background-color: white;
        }
            button.st-emotion-cache-13ejsyy.ef3psqc12{
            background-color: #2f9e36;
            color: #fff;
            transition: 0.2s;
            height: 2.5rem;
        }
        div.st-emotion-cache-1wmy9hl.e1f1d6gn0{
            width: 325px;
            height: 460px;
            background-color: #f0f0f0;
            border: 1px solid #ccc;
            margin: 20px;
            padding: 10px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            position: relative;
            left: 30px;
            border: 3px solid #73AD21;
            border-radius: 2rem;
            margin-top: 90px;
        }
        .st-bo{
            width: 300px;
        }
        .st-emotion-cache-10trblm{
            font-size: 25px;
            text-align: center;
            margin-right: 20px;
        }
        .st-emotion-cache-1vbkxwb p{
            font-size: 12px;
            text-align: center;
        }
        button.st-emotion-cache-7ym5gk.ef3psqc12{
           
            height: 1px;
        }
        .st-gw{
            height: auto;
            width: 300px;
        }
        .st-h9{
            
        }
        .dataframe tbody tr:nth-child(odd) {
            background-color: #f2f2f2;
        }
        .dataframe tbody tr:nth-child(even) {
            background-color: #dddddd;
        }
        table.dataframe {
            font-size: 20px;
            text-align: left;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

        set_background(background_image_path)
        email = st.text_input("Email Address")
        password = st.text_input("Password", type="password")

        login_button_clicked = st.button("LOGIN", key="login_button")
        reset_button_clicked = st.button("Forgot Password", key="reset_button")

        # Create a sign-up button with error handling
        if st.button("Sign Up"):
            if not validate_email(email):
                st.error("Please enter a valid email address.")
            elif not validate_password(password):
                st.error("Password must be at least 6 characters long.")
            else:
                signup(email, password)

        if login_button_clicked:
            if email and password:
                user = authenticate(email, password)
                if user:
                    st.session_state.user = user
                    st.success("Login successful.")
                    time.sleep(1)
            else:
                st.error("Please enter both email and password.")
                time.sleep(2)
        if reset_button_clicked:
            send_password_reset_email(email)


    else:
        st.sidebar.image("static/[Hi-Res] BAGONG PILIPINAS LOGO.png", use_column_width=True)

        st.sidebar.markdown("## :blue[Menu]")

        # Dropdown menu with options
        menu_selection = st.sidebar.selectbox(":blue[Select Option]", ["DATA 📖", "INVENTORY 📦", "MANAGE DATA ✍️", "FARMER 🌾", "Account 👤"])
        page_number = st.session_state.get('page_number', 1)

        if menu_selection == "DATA 📖":
            try:
                # Read the CSV file with a specified encoding and skip problematic lines
                df = pd.read_csv(csv_file_path, encoding='latin1')

                st.markdown("<h3>Search</h3>", unsafe_allow_html=True)
                search_query = st.text_input("", "")
                if search_query:
                    # Convert all columns to string type
                    df = df.astype(str)
                    # Filter the DataFrame based on the search query
                    df = df[df.apply(lambda row: row.str.contains(search_query, case=False)).any(axis=1)]
                
                # Calculate total number of rows and pages
                total_rows = len(df)
                num_pages = (total_rows + 9) // 10  # Ceiling division to get total number of pages

                # Input field to specify the page number
                page_number = st.number_input("Page Number", min_value=1, max_value=num_pages, value=page_number, key='page-number')

                # Calculate start and end indices for pagination
                start_index = (page_number - 1) * 10
                end_index = min(start_index + 10, total_rows)

                # Display the subset of rows for the specified page
                if total_rows > 0:
                    st.dataframe(df.iloc[start_index:end_index].reset_index(drop=True))
                else:
                    st.warning("No matching rows found.")
                
                # Display pagination information
                st.write(f"Showing page {page_number} of {num_pages}")

                # JavaScript for pagination
                pagination_script = """
                    <script>
                        function goToPage() {
                            var pageNumber = document.getElementById("page-number").value;
                            window.location.href = "?page=" + pageNumber;
                        }
                    </script>
                """
                st.markdown(pagination_script, unsafe_allow_html=True)              

            except Exception as e:
                st.error(f"Error reading CSV file: {e}")
                # If an exception occurs, retain the page number in session state
                st.session_state.page_number = page_number

        elif menu_selection == "MANAGE DATA ✍️":
            action = st.selectbox("Select Action", ["CREATE", "EDIT", "DELETE"])

            def authenticate_user(pin):
                if pin == "123456":  # Change the hardcoded PIN here
                    return True
                else:
                    return False
                

            if action == "CREATE":

                # Container for styling
                create_container = st.container()

                with create_container:
                    # Set the form width
                    st.markdown('<style>div.row-widget.stRadio > div{flex-direction:row;}</style>', unsafe_allow_html=True)
                    st.markdown('<style>div.row-widget.stRadio > div > label{flex-direction:row;}</style>', unsafe_allow_html=True)
                    st.markdown('<style>div.row-widget.stRadio > div > label > div{flex-direction:row;}</style>', unsafe_allow_html=True)
                    st.markdown('<style>div.row-widget.stRadio > div > label > span{margin-top: 25px;}</style>', unsafe_allow_html=True)

                    # Input fields for creating a new entry
                    name_of_cis = st.text_input("Name of CIS", placeholder="Enter CIS name")
                    location = st.text_input("Location", placeholder="Enter location")
                    source_of_water = st.text_input("Source of Water", placeholder="Enter water source")
                    scheme_of_irrigation = st.text_input("Scheme of Irrigation", placeholder="Enter irrigation scheme")
                    service_area = st.number_input("Service Area (Has.)", min_value=0.0, placeholder="Enter service area")
                    firmed_up_service_area = st.number_input("Firmed-Up Service Area (Has.)", min_value=0.0, placeholder="Enter firmed-up service area")
                    operational_area = st.number_input("Operational Area (Has.)", min_value=0.0, placeholder="Enter operational area")
                    no_of_farmer_beneficiaries = st.number_input("No. of Farmer Beneficiaries", min_value=0, placeholder="Enter number of farmer beneficiaries")
                    name_of_ia = st.text_input("Name of IA", placeholder="Enter IA name")
                    main_canals = st.text_input("Main Canal(s)", placeholder="Enter main canals")
                    laterals = st.text_input("Lateral(s) & Sub-Lateral(s)", placeholder="Enter laterals and sub-laterals")

                    # Create button to submit the form
                    create_button_clicked = st.button("Create Entry")

                    if create_button_clicked:
                        # Create a dictionary with the user inputs
                        new_entry = {
                            "Name of CIS": name_of_cis,
                            "Location": location,
                            "Source of Water": source_of_water,
                            "Scheme of Irrigation": scheme_of_irrigation,
                            "Service Area (Has.)": service_area,
                            "Firmed-Up Service Area (Has.)": firmed_up_service_area,
                            "Operational Area (Has.)": operational_area,
                            "No. of Farmer Beneficiaries": no_of_farmer_beneficiaries,
                            "Name of IA": name_of_ia,
                            "Main Canal(s)": main_canals,
                            "Lateral(s) & Sub-Lateral(s)": laterals
                        }
                        # Save the new entry to CSV file
                        save_to_csv(new_entry, csv_file_path)
                        st.success("Entry created successfully!")
                        # Print the new entry (you can replace this with your preferred way of saving the data)
                        st.write("New Entry:", new_entry)
                        
            elif action == "EDIT":                

                edit_container = st.container()
                with edit_container:
                    # Input field to specify the index of the entry to edit
                    edit_index = st.number_input("Index of Entry to Edit", min_value=0, placeholder="Enter index")
                    # Read the CSV file to retrieve data of the chosen index
                    try:
                        df = pd.read_csv(csv_file_path, encoding='latin1')
                        if edit_index < len(df):
                            # Display the data of the chosen index in input fields
                            name_of_cis = st.text_input("Name of CIS", value=df.loc[edit_index, "Name of CIS"])
                            location = st.text_input("Location", value=df.loc[edit_index, "Location"])
                            source_of_water = st.text_input("Source of Water", value=df.loc[edit_index, "Source of Water"])
                            scheme_of_irrigation = st.text_input("Scheme of Irrigation", value=df.loc[edit_index, "Scheme of Irrigation"])
                            service_area = st.number_input("Service Area (Has.)", value=df.loc[edit_index, "Service Area (Has.)"])
                            firmed_up_service_area = st.number_input("Firmed-Up Service Area (Has.)", value=df.loc[edit_index, "Firmed-Up Service Area (Has.)"])
                            operational_area = st.number_input("Operational Area (Has.)", value=df.loc[edit_index, "Operational Area (Has.)"])
                            no_of_farmer_beneficiaries = st.number_input("No. of Farmer Beneficiaries", value=df.loc[edit_index, "No. of Farmer Beneficiaries"])
                            name_of_ia = st.text_input("Name of IA", value=df.loc[edit_index, "Name of IA"])
                            main_canals = st.text_input("Main Canal(s)", value=df.loc[edit_index, "Main Canal(s)"])
                            laterals = st.text_input("Lateral(s) & Sub-Lateral(s)", value=df.loc[edit_index, "Lateral(s) & Sub-Lateral(s)"])

                            # Button to apply changes
                            apply_changes_button_clicked = st.button("Apply Changes")

                            if apply_changes_button_clicked:
                                # Create a dictionary with the edited entry
                                edited_entry = {
                                    "Name of CIS": name_of_cis,
                                    "Location": location,
                                    "Source of Water": source_of_water,
                                    "Scheme of Irrigation": scheme_of_irrigation,
                                    "Service Area (Has.)": service_area,
                                    "Firmed-Up Service Area (Has.)": firmed_up_service_area,
                                    "Operational Area (Has.)": operational_area,
                                    "No. of Farmer Beneficiaries": no_of_farmer_beneficiaries,
                                    "Name of IA": name_of_ia,
                                    "Main Canal(s)": main_canals,
                                    "Lateral(s) & Sub-Lateral(s)": laterals
                                }
                                # Apply changes to the entry at the specified index
                                edit_entry(edit_index, edited_entry, csv_file_path)
                        else:
                            st.warning("Index out of range.")
                    except Exception as e:
                        st.error(f"Error reading CSV file: {e}")

            elif action == "DELETE":
                
                delete_container = st.container()
                with delete_container:
                    # Read the CSV file
                    df = pd.read_csv(csv_file_path, encoding='latin1')

                    # Calculate total number of rows and pages
                    total_rows = len(df)
                    num_pages = (total_rows + 9) // 10  # Ceiling division to get total number of pages

                    # Input field to specify the page number
                    delete_page_number = st.number_input("Page Number", min_value=1, max_value=num_pages, value=1)

                    # Input field to specify the index of the entry to delete
                    delete_index = st.number_input("Index of Entry to Delete", min_value=0, placeholder="Enter index")

                    # Button to show the data of the entry
                    show_data_button_clicked = st.button("Show Data")

                    if show_data_button_clicked:
                        try:
                            # Calculate the actual index in the DataFrame based on the provided page number and index
                            actual_index = (delete_page_number - 1) * 10 + delete_index

                            # Check if the index is within the range of the DataFrame
                            if 0 <= actual_index < len(df):
                                # Show the data of the entry at the specified index in a table
                                st.write("Data of Entry:")
                                st.write(df.iloc[actual_index:actual_index+1])  # Displaying only the chosen entry as a DataFrame
                            else:
                                st.warning("Index out of range.")
                        except Exception as e:
                            st.error(f"Error displaying entry: {e}")

                    # Button to delete the entry
                    delete_button_clicked = st.button("Delete Entry")

                    if delete_button_clicked:
                        try:
                            # Calculate the actual index in the DataFrame based on the provided page number and index
                            actual_index = (delete_page_number - 1) * 10 + delete_index

                            # Check if the index is within the range of the DataFrame
                            if 0 <= actual_index < len(df):
                                # Delete the entry at the specified index
                                df = df.drop(actual_index)

                                # Save the modified DataFrame back to CSV
                                df.to_csv(csv_file_path, index=False)
                                st.success("Entry deleted successfully!")
                            else:
                                st.warning("Index out of range.")
                        except Exception as e:
                            st.error(f"Error deleting entry: {e}")


        elif menu_selection == "INVENTORY 📦":

            # Search functionality
            search_query = st.text_input("Search", key="inventory_search_input")

            if st.button("Search", key="inventory_search_button"):
                search_inventory(search_query)

            # Display the default inventory table if no search query is entered
            if not search_query:
                display_inventory(st.session_state.file_index)

            # Next button to cycle through files
            if st.button("Next"):
                st.session_state.file_index = (st.session_state.file_index + 1) % num_files


        elif menu_selection == "FARMER 🌾":
            st.markdown("TITLE")

            # Get the selected town from the dropdown
            farmer_menu_selection = st.selectbox(":green[SELECT TOWN]", list(town_file_paths.keys()))

            # Assuming farmer_menu_selection contains the selected town name
            selected_town = farmer_menu_selection

            # Check if the selected town is in the dictionary
            if selected_town in town_file_paths:
                # Get the file path for the selected town
                file_path = town_file_paths[selected_town]

                # Read the Excel file
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Display the DataFrame
                df = pd.read_excel(file_path)
                st.write(df)

                # Dropdown menu for actions
                action = st.selectbox("Select Action", ["Select Action", "Update", "Delete"])

                if action == "Update":
                    # Get the index and column to update
                    index = st.number_input("Enter the index of the row to update", min_value=2, max_value=sheet.max_row, step=1)

                    # Get the values from the selected row
                    row_values = [cell.value for cell in sheet[index]]

                    # Display the values
                    st.write("Current values:")
                    st.write(row_values)

                    # Input fields to edit the values
                    edited_values = []
                    for i, value in enumerate(row_values):
                        edited_value = st.text_input(f"Enter new value for {value}", value=value, key=f"{index}_{i}")
                        edited_values.append(edited_value)

                    if st.button("Apply Update"):
                        # Update the data in the Excel file
                        for col, value in zip(sheet[index], edited_values):
                            col.value = value

                        # Save the updated Excel file
                        wb.save(file_path)
                        st.write("Data updated successfully.")

                elif action == "DELETE ❌":
                    # Get the indices of the rows to delete
                    rows_to_delete = st.multiselect("Select rows to delete", df.index.tolist())

                    # Display the selected data before deletion
                    if rows_to_delete:
                        st.write("Selected data:")
                        st.write(df.loc[rows_to_delete])

                        if st.button("Delete Selected Rows"):
                            # Delete the selected rows from the DataFrame
                            df.drop(rows_to_delete, inplace=True)

                            # Update the Excel file
                            wb.remove(sheet)
                            sheet = wb.create_sheet(title='Sheet1', index=0)
                            for r_idx, row in enumerate(df.iterrows(), start=1):
                                for c_idx, value in enumerate(row[1], start=1):
                                    sheet.cell(row=r_idx, column=c_idx, value=value)

                            # Save the updated Excel file
                            wb.save(file_path)
                            st.write("Selected rows deleted successfully.")
                    else:
                        st.write("No rows selected for deletion.")


        elif menu_selection == "Account 👤":
                        st.subheader("MANAGE ACCOUNT")
                        try:
                            manage_accounts()
                        except Exception as e:
                            st.error(f"Error while managing accounts: {e}") 




        if st.session_state.user is not None:
                # Add logout button
            if st.sidebar.button("Logout"):
                logout()




if __name__ == "__main__":
    main()